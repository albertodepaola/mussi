#!/usr/bin/python
# -*- coding: utf-8 -*-

from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import fills, PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from django.template import loader
from django.core.urlresolvers import reverse
import random
import string
import re
import os
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

linked_prop_email = 'linkedprop@gmail.com'
linked_prop_telephone = '+54 9 11 1111 2222'


def merge_two_dicts(x, y):
    z = x.copy()
    z.update(y)
    return z


def create_html_plain_email(subject, plain_body, html_body, to, attachments_paths=[]):
    msg = EmailMultiAlternatives(subject, plain_body, to=to)
    msg.attach_alternative(html_body, 'text/html')
    for attachment in attachments_paths:
        msg.attach_file(attachment['file_path'], attachment['mime_type'])

    return msg


def send_welcome_email(request, user, password):
    template = loader.get_template('collaborative_crm/email_welcome.html')
    template_plain = loader.get_template('collaborative_crm/email_welcome.txt')
    context = {
        'created_user': user,
        'password': password,
        'link_to_login': request.build_absolute_uri(reverse('login_page'))
    }
    create_html_plain_email(u'[LinkedProp] Bienvenido', template_plain.render(context, request),
                            template.render(context, request), [user.username]).send()


def new_dir_name():
    dir_name = settings.TEMP_URL + ''.join(random.choice(string.digits + string.ascii_letters) for _ in range(16))
    while os.path.isdir(dir_name):
        dir_name = settings.TEMP_URL + ''.join(random.choice(string.digits + string.ascii_letters) for _ in range(16))
    os.makedirs(dir_name)
    return dir_name


def number_to_excel_col(number):
    col = ''
    if number > 26:
        char_num = (number - 1) // 26
        col = chr(64 + char_num)
        number -= 26 * char_num

    return col + chr(64 + number)


def parse_format_string(format_string):
    return {
        'bold': 'bold' in format_string,
        '$0': '$0' in format_string,
        'wrap_text': 'wrap-text' in format_string,
        'v_center': 'v-center' in format_string,
        'h_center': 'h-center' in format_string,
        'border': {
            'top': ('thick' if 'border-top-thick' in format_string else 'thin')
            if 'border-top-' in format_string else None,
            'left': ('thick' if 'border-left-thick' in format_string else 'thin')
            if 'border-left-' in format_string else None,
            'right': ('thick' if 'border-right-thick' in format_string else 'thin')
            if 'border-right-' in format_string else None,
            'bottom': ('thick' if 'border-bottom-thick' in format_string else 'thin')
            if 'border-bottom-' in format_string else None
        } if 'border-' in format_string else None,
        'row_height': int(re.search('row-height-([0-9]+)', format_string).group(1))
        if re.search('row-height-([0-9]+)', format_string) else None,
        'col_width': int(re.search('col-width-([0-9]+)', format_string).group(1))
        if re.search('col-width-([0-9]+)', format_string) else None,
        'font_size': int(re.search('font-size-([0-9]+)', format_string).group(1))
        if re.search('font-size-([0-9]+)', format_string) else None
    }


def apply_cell_format(cell, format_dict):
    if format_dict['$0']:
        cell.number_format = '$0'
    cell.alignment = Alignment(wrapText=format_dict['wrap_text'],
                               vertical='center' if format_dict['v_center'] else None,
                               horizontal='center' if format_dict['h_center'] else None)


def generate_excel(title, values, cover_image={}, images={}, hide_gridlines=False):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title

    if hide_gridlines:
        worksheet.sheet_view.showGridLines = False

    for val in values:
        if val.get('condition', True):
            col_from = val['col'] if val['col'].__class__ == int else val['col'][0]
            col_to = None if val['col'].__class__ == int else val['col'][1]
            row_from = val['row'] if val['row'].__class__ == int else val['row'][0]
            row_to = None if val['row'].__class__ == int else val['row'][1]
            format_dict = parse_format_string(val['format'])

            cell = worksheet.cell(column=col_from, row=row_from, value=val['value'])
            apply_cell_format(cell, format_dict)
            if format_dict['font_size']:
                if col_to or row_to:
                    for cell_font_size in [{'col': col, 'row': row}
                                           for col in range(col_from, (col_to if col_to else col_from) + 1)
                                           for row in range(row_from, (row_to if row_to else row_from) + 1)]:
                        worksheet.cell(column=cell_font_size['col'], row=cell_font_size['row']).font = \
                            Font(size=format_dict['font_size'], bold=format_dict['bold'])
                else:
                    cell.font = Font(size=format_dict['font_size'], bold=format_dict['bold'])
            if format_dict['border']:
                kwargs = {}
                for border in [border for border in format_dict['border'].keys() if format_dict['border'][border]]:
                    kwargs[border] = Side(style=format_dict['border'][border])
                if col_to or row_to:
                    for cell_border in [{'col': col, 'row': row}
                                        for col in range(col_from, (col_to if col_to else col_from) + 1)
                                        for row in range(row_from, (row_to if row_to else row_from) + 1)]:
                        worksheet.cell(column=cell_border['col'], row=cell_border['row']).border = Border(**kwargs)
                else:
                    cell.border = Border(**kwargs)
            if format_dict['row_height']:
                worksheet.row_dimensions[row_from].height = format_dict['row_height']
            if format_dict['col_width']:
                worksheet.column_dimensions[number_to_excel_col(col_from)].width = format_dict['col_width']

            if col_to or row_to:
                worksheet.merge_cells(start_row=row_from, start_column=col_from,
                                      end_row=row_to if row_to else row_from,
                                      end_column=col_to if col_to else col_from)

    if cover_image and cover_image['url']:
        img = Image(cover_image['url'], size=(300, 300), nochangeaspect=False)
        img.anchor(worksheet.cell(cover_image['cell']))
        worksheet.add_image(img)

    if images and images['images']:
        row_num = images['start_at_row']
        cell = worksheet.cell(column=images['col'], row=row_num, value=u'Imágenes')
        cell.font = Font(bold=True)
        row_num += 2
        image_count = 0
        for image in images['images']:
            if image.description:
                cell = worksheet.cell(column=images['col'], row=row_num, value=image.description.replace(u'\n', u' '))
                cell.font = Font(bold=True)
                row_num += 2
            img = Image(image.image_path)
            # se resta image_count porque por algun motivo las imagenes se defasan a mayor numero de row (probablemente
            # un error en el calculo de la altura de las filas por la libreria)
            img.anchor(worksheet.cell('{0}{1}'.format(number_to_excel_col(images['col']),
                                                      row_num - image_count - (image_count // 4))))
            if img.drawing.width > 800:
                img.drawing.width = 800
            if img.drawing.height > 600:
                img.drawing.height = 600
            worksheet.add_image(img)
            row_num += (img.drawing.height // 18)
            image_count += 1

    file_name = '{0}/{1}.xlsx'.format(new_dir_name(), slugify(title))
    workbook.save(file_name)
    return file_name


pdf_default_row_height = cm / 2
pdf_default_col_width = 10.27
styles = getSampleStyleSheet()


def calculate_top(top, elem_height=0):
    return - top - elem_height


def excel_to_pdf_width(excel_width):
    return excel_width * (cm / 5.5)


def calculate_left_from_col_num(col_num, col_widths):
    return sum(col_widths[:col_num - 1])


def calculate_top_from_row_num(row_num, row_heights, elem_height=0, min_row=0, max_row=0):
    return calculate_top(sum(row_heights[:row_num - 1]),
                         elem_height if elem_height else sum(row_heights[min_row - 1:max_row]))


def calculate_image_dimension(dimension):
    return dimension * (cm / 60)


def generate_pdf(title, values, cover_image={}, images={}):
    file_name = '{0}/{1}.pdf'.format(new_dir_name(), slugify(title))
    page_size = landscape(A4)
    margin = cm
    canvas = Canvas(file_name, pagesize=page_size)
    page_width, page_height = page_size
    canvas.translate(margin, page_height - margin)

    col_widths = [{'col': v['col'], 'width': parse_format_string(v['format'])['col_width']} for v in values if
                  parse_format_string(v['format'])['col_width']]
    for col_num in range(1, max([v['col'] if v['col'].__class__ == int else v['col'][1] for v in values])):
        if not [c for c in col_widths if c['col'] == col_num]:
            col_widths.append({'col': col_num, 'width': pdf_default_col_width })
    col_widths = [excel_to_pdf_width(col['width']) for col in sorted(col_widths, key=lambda c: c['col'])]

    row_heights = [{'row': v['row'], 'height': parse_format_string(v['format'])['row_height']} for v in values if
                   parse_format_string(v['format'])['row_height']]
    for row_num in range(1, max([v['row'] if v['row'].__class__ == int else v['row'][1] for v in values])):
        if not [r for r in row_heights if r['row'] == row_num]:
            row_heights.append({'row': row_num, 'height': pdf_default_row_height})
    row_heights = [row['height'] for row in sorted(row_heights, key=lambda r: r['row'])]

    canvas.setFont('Helvetica-Bold', 14)
    canvas.drawString(0, 0, title)

    if cover_image and cover_image['url']:
        cover_image_width_height = 6*cm
        canvas.drawImage(cover_image['url'], 0, calculate_top(cm, cover_image_width_height), cover_image_width_height,
                         cover_image_width_height)

    for table in list({value['table'] for value in values if 'table' in value.keys()}):
        table_values = [v for v in values if 'table' in v.keys() and v['table'] == table]
        min_row = min([v['row'] if v['row'].__class__ == int else v['row'][0] for v in table_values])
        max_row = max([v['row'] if v['row'].__class__ == int else v['row'][1] for v in table_values])
        min_col = min([v['col'] if v['col'].__class__ == int else v['col'][0] for v in table_values])
        max_col = max([v['col'] if v['col'].__class__ == int else v['col'][1] for v in table_values])
        data = []
        table_style = [('FONTSIZE', (0, 0), (-1, -1), 8.5)]

        for row in sorted(list({v['row'] for v in table_values})):
            row_list = []
            for value in sorted([v for v in table_values if v['row'] == row],
                                key=lambda v: v['col'] if v['col'].__class__ == int else v['col'][0]):

                value_min_row = (value['row'] if value['row'].__class__ == int else value['row'][0]) - min_row
                value_max_row = (value['row'] if value['row'].__class__ == int else value['row'][1]) - min_row
                value_min_col = (value['col'] if value['col'].__class__ == int else value['col'][0]) - min_col
                value_max_col = (value['col'] if value['col'].__class__ == int else value['col'][1]) - min_col
                format_dict = parse_format_string(value['format'])
                if value_min_row != value_max_row or value_min_col != value_max_col:
                    table_style.append(('SPAN', (value_min_col, value_min_row), (value_max_col, value_max_row)))
                if format_dict['bold']:
                    table_style.append(('FONT', (value_min_col, value_min_row), (value_max_col, value_max_row),
                                        'Helvetica-Bold'))
                if format_dict['h_center']:
                    table_style.append(('ALIGNMENT', (value_min_col, value_min_row), (value_max_col, value_max_row),
                                        'CENTER'))
                if format_dict['border']:
                    if format_dict['border']['top']:
                        table_style.append(('LINEABOVE', (value_min_col, value_min_row),
                                            (value_max_col, value_max_row),
                                            1 if format_dict['border']['top'] == 'thin' else 2, colors.black))
                    if format_dict['border']['left']:
                        table_style.append(('LINEBEFORE', (value_min_col, value_min_row),
                                            (value_max_col, value_max_row),
                                            1 if format_dict['border']['left'] == 'thin' else 2, colors.black))
                    if format_dict['border']['right']:
                        table_style.append(('LINEAFTER', (value_min_col, value_min_row),
                                            (value_max_col, value_max_row),
                                            1 if format_dict['border']['right'] == 'thin' else 2, colors.black))
                    if format_dict['border']['bottom']:
                        table_style.append(('LINEBELOW', (value_min_col, value_min_row),
                                            (value_max_col, value_max_row),
                                            1 if format_dict['border']['bottom'] == 'thin' else 2, colors.black))

                if not format_dict['wrap_text']:
                    cell_value = value['value']
                else:
                    paragraph_text = unicode(value['value']).replace(u'\n', u'<br/>')
                    break_at = 0
                    max_chars_per_line = 105
                    max_lines = 4
                    lines = 0
                    line_num = 0
                    line_begin_at = 0
                    for line in paragraph_text.split(u'<br/>'):
                        line_size = (len(line) // max_chars_per_line) + 1
                        lines += line_size
                        line_num += 1
                        if lines >= max_lines:
                            break_at = sum([len(l) for l in paragraph_text.split(u'<br/>')[:line_num]]) + \
                                       5 * (line_num - 1) - (len(line) - (((max_lines - line_begin_at) * max_chars_per_line)) if len(line) > (max_lines - line_begin_at) * max_chars_per_line else 0)
                            break
                        line_begin_at = lines

                    paragraph_text = paragraph_text[:break_at if break_at else len(paragraph_text)]
                    while paragraph_text[-5:] == u'<br/>':
                        paragraph_text = paragraph_text[:-5]

                    cell_value = Paragraph(paragraph_text + (u'...' if break_at else u''), styles['Normal'])
                if value['col'].__class__ == int:
                    row_list.append(cell_value)
                else:
                    for col in range(*value['col']) + [value['col'][1]]:
                        row_list.append(cell_value if col == value['col'][0] else '')
            data.append(row_list)
        canvas.setFont('Helvetica', 8)
        table = Table(data, colWidths=col_widths[min_col - 1:max_col], rowHeights=row_heights[min_row - 1:max_row],
                      style=table_style)
        table.wrapOn(canvas, page_width, page_height)
        table.drawOn(canvas, calculate_left_from_col_num(min_col, col_widths),
                     calculate_top_from_row_num(min_row, row_heights, min_row=min_row, max_row=max_row))

    canvas.showPage()
    canvas.translate(margin, page_height - margin)

    if images and images['images']:
        image_count = 0
        for image in images['images']:
            if image.description:
                canvas.setFont('Helvetica-Bold', 11)
                canvas.drawString((page_width - margin * 2) * 0.5 * (image_count % 2), 0, image.description.replace(u'\n', u' '))
            img = Image(image.image_path)
            if img.drawing.width > 800:
                img.drawing.width = 800
            if img.drawing.height > 1000:
                img.drawing.height = 1000
            canvas.drawImage(image.image_path, (page_width - margin * 2) * 0.5 * (image_count % 2),
                             calculate_top(cm, calculate_image_dimension(img.drawing.height)),
                             calculate_image_dimension(img.drawing.width),
                             calculate_image_dimension(img.drawing.height))
            if image_count % 2:
                canvas.showPage()
                if image != images['images'][-1:][0]:
                    canvas.translate(margin, page_height - margin)
            image_count += 1

    canvas.save()

    return file_name
